"""
maricopa_nts_parser_v01.py
Parses grouped NTS PDFs → structured Excel/CSV output.
Reads PDFs from grouped_output/pdfs/, writes to parsed_output/.

Fields extracted:
  Doc Number, First Name, Last Name, 2nd First, 2nd Last,
  Street Address, City, State, Postal Code,
  Property Address, Property City, Property State, Property Postal Code,
  County, Parcel Number, Original Loan, Trustee Name, Trustee Phone,
  Auction Date, PDF Path

Dependencies:
  pip install pdfplumber openpyxl
"""

import logging
import re
import sys
from pathlib import Path

log = logging.getLogger("nts_parser")
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

PDF_DIR     = Path("grouped_output/pdfs")
OUTPUT_DIR  = Path("parsed_output")

HEADERS = [
    "Doc Number", "First Name", "Last Name", "2nd First", "2nd Last",
    "Street Address", "City", "State", "Postal Code",
    "Property Address", "Property City", "Property State", "Property Postal Code",
    "County", "Parcel Number", "Original Loan",
    "Trustee Name", "Trustee Phone", "Auction Date", "PDF Path",
]

SUFFIXES = {"JR", "SR", "II", "III", "IV", "TRUST", "LLC", "CORP", "INC", "LP", "LLP"}


def run():
    try:
        import pdfplumber
        import openpyxl
    except ImportError:
        log.error("Missing deps: pip install pdfplumber openpyxl")
        sys.exit(1)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    pdfs = sorted(PDF_DIR.glob("*.pdf"))
    log.info(f"Found {len(pdfs)} PDFs in {PDF_DIR}")

    if not pdfs:
        log.warning("No PDFs found. Run maricopa_nts_png_v01.py first.")
        return

    records = []
    for pdf_path in pdfs:
        log.info(f"Parsing: {pdf_path.name}")
        try:
            text = _extract_text(pdf_path)
            rec  = _parse_nts_text(text, pdf_path)
            records.append(rec)
            log.info(f"  → {rec.get('Doc Number')} | {rec.get('First Name')} {rec.get('Last Name')} | {rec.get('Auction Date')}")
        except Exception as exc:
            log.error(f"  Parse failed: {exc}", exc_info=True)
            records.append({"Doc Number": pdf_path.stem, "PDF Path": str(pdf_path)})

    # ── write outputs ──────────────────────────────────────────────────────
    _write_excel(records)
    _write_csv(records)
    log.info(f"Done. {len(records)} records parsed → {OUTPUT_DIR}")


def _extract_text(pdf_path: Path) -> str:
    import pdfplumber
    all_text = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            try:
                t = page.extract_text()
                if t:
                    all_text.append(t)
            except Exception:
                pass
    return "\n".join(all_text)


def _parse_nts_text(text: str, pdf_path: Path) -> dict:
    rec = {h: "" for h in HEADERS}
    rec["PDF Path"] = str(pdf_path)
    rec["County"]   = "Maricopa"

    lines = [l.strip() for l in text.splitlines() if l.strip()]

    # ── doc number ─────────────────────────────────────────────────────────
    for line in lines[:20]:
        m = re.search(r"\b(20\d{2}\d{7,9})\b", line)
        if m:
            rec["Doc Number"] = m.group(1)
            break
    if not rec["Doc Number"]:
        rec["Doc Number"] = pdf_path.stem

    # ── trustor / owner name ───────────────────────────────────────────────
    name_raw = _find_after(text, [
        r"[Tt]rustor[:\s]+",
        r"[Gg]rantor[:\s]+",
        r"[Oo]bligor[:\s]+",
        r"[Bb]orrower[:\s]+",
    ])
    if name_raw:
        f1, l1, f2, l2 = _parse_name(name_raw)
        rec["First Name"] = f1
        rec["Last Name"]  = l1
        rec["2nd First"]  = f2
        rec["2nd Last"]   = l2

    # ── mailing address ────────────────────────────────────────────────────
    mail_raw = _find_after(text, [
        r"[Mm]ailing\s+[Aa]ddress[:\s]+",
        r"[Tt]rustor.*?[Aa]ddress[:\s]+",
    ])
    if mail_raw:
        addr = _parse_address(mail_raw)
        if addr:
            rec["Street Address"] = addr.get("street", "")
            rec["City"]           = addr.get("city", "")
            rec["State"]          = addr.get("state", "")
            rec["Postal Code"]    = addr.get("zip", "")

    # ── property address ───────────────────────────────────────────────────
    prop_raw = _find_after(text, [
        r"[Pp]roperty\s+[Aa]ddress[:\s]+",
        r"[Ss]itus[:\s]+",
        r"[Pp]remises[:\s]+",
        r"[Ll]ocated at[:\s]+",
    ])
    if prop_raw:
        addr = _parse_address(prop_raw)
        if addr:
            rec["Property Address"]      = addr.get("street", "")
            rec["Property City"]         = addr.get("city", "")
            rec["Property State"]        = addr.get("state", "AZ")
            rec["Property Postal Code"]  = addr.get("zip", "")

    # ── parcel ─────────────────────────────────────────────────────────────
    m = re.search(r"(?:APN|[Pp]arcel)[:\s#]+(\d{3}[-\s]\d{2}[-\s]\d{3}[A-Z]?)", text)
    if m:
        rec["Parcel Number"] = m.group(1).replace(" ", "-")

    # ── original loan ──────────────────────────────────────────────────────
    m = re.search(
        r"(?:[Oo]riginal\s+[Ll]oan|[Uu]npaid\s+[Bb]alance|[Ll]oan\s+[Aa]mount)[:\s]+\$?([\d,]+(?:\.\d{2})?)",
        text,
    )
    if m:
        rec["Original Loan"] = "$" + m.group(1)

    # ── trustee ────────────────────────────────────────────────────────────
    trustee_raw = _find_after(text, [r"[Tt]rustee[:\s]+", r"[Ss]ubstitute\s+[Tt]rustee[:\s]+"])
    if trustee_raw:
        rec["Trustee Name"] = trustee_raw[:80]

    m_phone = re.search(r"(?:Phone|Tel|Ph)[:\s]*\(?\d{3}\)?[\s\-\.]\d{3}[\s\-\.]\d{4}", text)
    if m_phone:
        digits = re.search(r"(\d{3})[\)\s\-\.]+(\d{3})[\s\-\.](\d{4})", m_phone.group(0))
        if digits:
            rec["Trustee Phone"] = f"({digits.group(1)}) {digits.group(2)}-{digits.group(3)}"

    # ── auction date ───────────────────────────────────────────────────────
    date_patterns = [
        r"[Ss]ale\s+[Dd]ate[:\s]+(\w+\s+\d{1,2},?\s+\d{4})",
        r"[Aa]uction\s+[Dd]ate[:\s]+(\w+\s+\d{1,2},?\s+\d{4})",
        r"[Ss]old\s+(?:on|at)[:\s]+(\w+\s+\d{1,2},?\s+\d{4})",
        r"(\w+\s+\d{1,2},\s+\d{4})(?=.*[Ss]ale)",
    ]
    for pat in date_patterns:
        m = re.search(pat, text)
        if m:
            raw_d = m.group(1).strip()
            rec["Auction Date"] = _norm_date(raw_d)
            break

    return rec


# ── helpers ────────────────────────────────────────────────────────────────────
def _find_after(text: str, patterns: list) -> str:
    """Find the first match of any pattern, return text after the match."""
    for pat in patterns:
        m = re.search(pat, text)
        if m:
            remainder = text[m.end():].strip()
            # Take up to first blank line or 200 chars
            snippet = remainder.split("\n\n")[0][:200]
            return snippet.split("\n")[0].strip()
    return ""


def _parse_name(raw: str) -> tuple:
    """Returns (first1, last1, first2, last2)."""
    parts = re.split(r"\s+(?:&|AND)\s+", raw.upper(), maxsplit=1)

    def split_one(s):
        tokens = s.strip().split()
        while tokens and tokens[-1] in SUFFIXES:
            tokens.pop()
        if not tokens:
            return "", ""
        if len(tokens) == 1:
            return "", tokens[0].title()
        return " ".join(tokens[1:]).title(), tokens[0].title()

    f1, l1 = split_one(parts[0])
    f2, l2 = split_one(parts[1]) if len(parts) > 1 else ("", "")
    return f1, l1, f2, l2


def _parse_address(raw: str) -> dict | None:
    raw = " ".join(raw.split())
    m = re.search(
        r"(\d+\s+[A-Za-z0-9\s\.#,\-]+?),?\s+([A-Za-z\s]+),?\s+([A-Z]{2})\s+(\d{5}(?:-\d{4})?)",
        raw,
    )
    if m:
        return {
            "street": m.group(1).strip(",").strip(),
            "city":   m.group(2).strip(",").strip(),
            "state":  m.group(3).strip(),
            "zip":    m.group(4).strip(),
        }
    return None


def _norm_date(raw: str) -> str:
    from datetime import datetime
    for fmt in ("%B %d, %Y", "%b %d, %Y", "%B %d %Y", "%m/%d/%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(raw.strip(), fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return raw


def _write_excel(records: list):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NTS Leads"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_align = Alignment(horizontal="center")

    for col_i, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_i, value=header)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align

    for row_i, rec in enumerate(records, 2):
        for col_i, header in enumerate(HEADERS, 1):
            ws.cell(row=row_i, column=col_i, value=rec.get(header, ""))

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    path = OUTPUT_DIR / "nts_data.xlsx"
    wb.save(path)
    log.info(f"Excel → {path}")


def _write_csv(records: list):
    import csv
    path = OUTPUT_DIR / "nts_data.csv"
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=HEADERS, extrasaction="ignore")
        writer.writeheader()
        for rec in records:
            writer.writerow({h: rec.get(h, "") for h in HEADERS})
    log.info(f"CSV → {path}")


if __name__ == "__main__":
    run()
